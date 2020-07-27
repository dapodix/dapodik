import logging
from dapodik import Dapodik
from dataclasses import asdict
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
from random import randint
from dapodik.peserta_didik import PesertaDidik, PesertaDidikBaru, PesertaDidikLongitudinal, RegistrasiPesertaDidik


def get_data(ws: Worksheet, row: int):
    def get_value(col):
        data = ws[f"{col}{row}"].value
        if not data:
            return data
        if type(data) == str or type(data) == int:
            return data
        elif type(data) == datetime:
            return data.strftime("%Y-%m-%d")

    data = {}
    data['nama_pd'] = get_value('B')
    data['nama'] = get_value('B')
    data['jenis_kelamin'] = get_value('C')
    data['nik'] = get_value('D')
    data['tempat_lahir'] = get_value('E')
    data['tanggal_lahir'] = get_value('F')
    data['no_kk'] = get_value('G')
    data['reg_akta_lahir'] = get_value('H')
    data['agama_id'] = get_value('AG')
    data['alamat_jalan'] = get_value('J')
    data['rt'] = get_value('K')
    data['rw'] = get_value('L')
    data['nama_dusun'] = get_value('M')
    data['desa_kelurahan'] = get_value('N')
    #data['kode_wilayah'] = get_value('N')
    data['kode_pos'] = get_value('O')
    data['anak_keberapa'] = get_value('P')
    data['nama_ayah'] = get_value('Q')
    data['nik_ayah'] = get_value('R')
    data['tahun_lahir_ayah'] = get_value('AH')
    data['jenjang_pendidikan_ayah'] = get_value('AI')
    data['pekerjaan_id_ayah'] = get_value('AJ')
    data['penghasilan_id_ayah'] = get_value('AK')
    data['nama_ibu_kandung'] = get_value('V')
    data['nik_ibu'] = get_value('W')
    data['tahun_lahir_ibu'] = get_value('AL')
    data['jenjang_pendidikan_ibu'] = get_value('AM')
    data['pekerjaan_id_ibu'] = get_value('AN')
    data['penghasilan_id_ibu'] = get_value('AO')
    data['menit_tempuh_ke_sekolah'] = randint(1, 10)
    # Longitudinal
    tinggi_badan = get_value('AB')
    data['tinggi_badan'] = tinggi_badan if tinggi_badan else randint(98, 105)
    berat_badan = get_value('AC')
    data['berat_badan'] = berat_badan if berat_badan else randint(12, 16)
    lingkar_kepala = get_value('AD')
    data['lingkar_kepala'] = lingkar_kepala if lingkar_kepala else randint(
        19, 22)
    # Registrasi
    data['nipd'] = get_value('A')
    jarak_rumah = randint(1, 7)
    if jarak_rumah > 1:
        data['jarak_rumah_ke_sekolah'] = 2
        data['jarak_rumah_ke_sekolah_km'] = jarak_rumah
    return data


def import_pd_dari_excel(d: Dapodik, filename: str, start: int, end: int, default_pdb: dict = {}, default_pd: dict = {}, default_reg: dict = {}):
    wb = load_workbook(filename, data_only=True)
    ws = wb.active
    sekolah_id = d.Sekolah()[0].sekolah_id
    default = {
        'sekolah_id': sekolah_id
    }
    for row in range(start, end):
        data: dict = get_data(ws, row)
        data['sekolah_id'] = sekolah_id
        # data.update(default)
        # peserta didik baru
        pdb_new: PesertaDidikBaru = d.PesertaDidikBaru.new(
            data, default=default, params={
                'sekolah_id': sekolah_id
            })
        data.update(asdict(pdb_new))
        # peserta didk=ik
        data.update(default_pd)
        pd_new: PesertaDidik = d.PesertaDidik.new(
            data, default=default, params={
                'sekolah_id': sekolah_id,
                'pd_module': 'pdterdaftar'
            })
        logging.info(f"Berhasil dimasukan {pd_new.nama}")
        # data longitudinal
        d.post(d._url+'customrest/savePdb', json={
            'pdb_id': data.get('pdb_id')
        })
        data.update(default_reg)
        data['peserta_didik_id'] = pd_new.peserta_didik_id
        pdl_new: PesertaDidikLongitudinal = d.PesertaDidikLongitudinal.new(
            data, default=default, params=default)
        logging.info(f"Berhasil longitudinal {pd_new.nama}")
        # data registrasi
        #pdr = RegistrasiPesertaDidik(**data)
        pdr_new: RegistrasiPesertaDidik = d.RegistrasiPesertaDidik.new(
            data, default=default)
        logging.info(f"Berhasil registrasi {pd_new.nama}")
